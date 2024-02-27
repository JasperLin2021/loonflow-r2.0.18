from _decimal import Decimal
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font

def convert_to_decimal(value):
    try:
        return Decimal(value)
    except (TypeError, ValueError):
        raise ValueError('initial_balance_amount must be a number')
def add_hyperlink(sheet, cell, link, display_text):
    col = 0
    for l in link:
       # 创建一个超链接对象
        sheet.cell(row=cell.row, column=cell.column + col).hyperlink = l
        # # 设置单元格样式
        sheet.cell(row=cell.row, column=cell.column + col).font = Font(underline="single", color="0563C1")
        # # 设置显示文本
        sheet.cell(row=cell.row, column=cell.column + col).value = display_text + str(col + 1)
        col = col + 1

def get_column_letter(column_index):
    base = ord('A')
    letters = []
    while column_index > 0:
        column_index, remainder = divmod(column_index - 1, 26)
        letters.append(chr(base + remainder))
    letters.reverse()
    return ''.join(letters)
