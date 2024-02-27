import json
from datetime import datetime


from django.http import HttpResponse
from openpyxl.drawing.text import Hyperlink
from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from schema import Schema, And, Optional, Use

from apps.loon_base_view import LoonBaseView
from apps.ticket.models import TicketRecord
from extensions.utils import add_hyperlink, get_column_letter, convert_to_decimal
from service.finance.finance_base_service import finance_base_service_ins
from service.format_response import api_response
from service.ticket.ticket_base_service import ticket_base_service_ins




# Create your views here.
class LoonAccountView(LoonBaseView):
    # 用于验证POST请求数据的模式
    post_schema = Schema({
        'holder': And(str, lambda n: n != ''),
        'card_number': And(str, lambda n: n != ''),
        'card_bank': And(str, lambda n: n != ''),
        'initial_balance_amount': And(Use(convert_to_decimal), lambda n: n != ''),
        Optional('description'): str,
    })


    def get(self, request, *args, **kwargs):
        """
        用户角色列表
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        request_data = request.GET
        search_value = request_data.get('search_value', '')
        per_page = int(request_data.get('per_page', 10))
        page = int(request_data.get('page', 1))
        flag, result = finance_base_service_ins.get_account_list(search_value, page, per_page)
        if flag is not False:
            data = dict(value=result.get('account_result_object_format_list'),
                        per_page=result.get('paginator_info').get('per_page'),
                        page=result.get('paginator_info').get('page'),
                        total=result.get('paginator_info').get('total'))
            code, msg, = 0, ''
        else:
            code, data = -1, ''
        return api_response(code, msg, data)

    def post(self, request, *args, **kwargs):
        """
        add role
        新增角色
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        json_str = request.body.decode('utf-8')
        request_data_dict = json.loads(json_str)
        holder = request_data_dict.get('holder')
        card_number = request_data_dict.get('card_number')
        card_bank = request_data_dict.get('card_bank')
        description = request_data_dict.get('description')
        initial_balance_amount = balance_amount = request_data_dict.get('initial_balance_amount')
        modifier = creator = request.user.username

        flag, result = finance_base_service_ins.add_account(holder=holder, card_number=card_number, card_bank=card_bank, initial_balance_amount=initial_balance_amount,
                                                         balance_amount=balance_amount, creator=creator, modifier=modifier, description=description)
        if flag is False:
            return api_response(-1, result, {})
        return api_response(0, result, {})


class LoonAccountDetailView(LoonBaseView):
    post_schema = Schema({
        'account_id': And(str, lambda n: n != ''),
        Optional('description'): str,
    })

    def patch(self, request, *args, **kwargs):
        """
        update role
        更新角色信息
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        account_id = kwargs.get('account_id')
        json_str = request.body.decode('utf-8')
        request_data_dict = json.loads(json_str)
        description = request_data_dict.get('description')
        modifier = request.user.username
        gmt_modified = datetime.now()
        flag, result = finance_base_service_ins.update_account(account_id, description, modifier, gmt_modified)
        if flag is False:
            return api_response(-1, result, {})
        return api_response(0, '', {})

    def delete(self, request, *args, **kwargs):
        """
        delete role
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        account_id = kwargs.get('account_id')
        flag, result = finance_base_service_ins.delete_account(account_id)
        if flag is False:
            return api_response(-1, result, {})
        return api_response(0, '', {})


class LoonCapitalFlowView(LoonBaseView):
    def get(self, request, *args, **kwargs):
        """
        用户角色列表
        :param request:
        :param args:
        :param kwargs:
        :return:
        """

        request_data = request.GET

        # username = request.META.get('HTTP_USERNAME')
        card_bank = request_data.get('card_bank', '')
        card_number = request_data.get('card_number', '')
        capital_flow_type = request_data.get('capital_flow_type')
        create_start = request_data.get('create_start', '')
        create_end = request_data.get('create_end', '')
        reverse = int(request_data.get('reverse', 1))
        per_page = int(request_data.get('per_page', 10))
        page = int(request_data.get('page', 1))

        # 未指定创建起止时间则取最近三年的记录
        if not(create_start or create_end):
            import datetime
            end_time = datetime.datetime.now() + datetime.timedelta(hours=1)
            last_year_time = datetime.datetime.now() - datetime.timedelta(days=7)
            create_start = str(last_year_time)[:19]
            create_end = str(end_time)[:19]


        flag, result = finance_base_service_ins.get_capital_flow_list(card_bank=card_bank, card_number=card_number, capital_flow_type=capital_flow_type,
                                                                      create_start=create_start, create_end=create_end, reverse=reverse,
                                                                      per_page=per_page, page=page)
        if flag is not False:
            data = dict(value=result.get('capital_flow_result_object_format_list'),
                        per_page=result.get('paginator_info').get('per_page'),
                        page=result.get('paginator_info').get('page'),
                        total=result.get('paginator_info').get('total'))
            code, msg, = 0, ''
        else:
            code, data = -1, ''
        return api_response(code, msg, data)


class LoonCapitalFlowExportView(LoonBaseView):
    def get(self, request, *args, **kwargs):
        username = request.META.get('HTTP_USERNAME')
        result = LoonCapitalFlowView.get(self, request, *args, **kwargs)
        # 将字节数据转换为字符串
        data_str = result.content.decode('utf-8')
        # 解析JSON字符串
        json_data = json.loads(data_str)

        export_list = []

        for item in json_data['data']['value']:

            ticket_id = item['ticket_record_id']
            ticket_detail = ticket_base_service_ins.get_ticket_detail(ticket_id, username)
            ticket_detail_field_list = ticket_detail[1]['field_list']
            ticket_obj = TicketRecord.objects.filter(id=ticket_id, is_deleted=False).first()
            item['流水号'] = ticket_obj.sn

            ticket_detail_format = dict()
            for key in ticket_detail_field_list:
                if key['field_name'] == "附件上传" and key['field_value'] != '[]':
                    url_list = []
                    for u in json.loads(key['field_value']):
                        url_list.append("http://127.0.0.1:8001" + u.get("url"))
                    ticket_detail_format[key['field_name']] = url_list
                elif key['field_name'] == "当前处理人":
                    continue
                else:
                    ticket_detail_format[key['field_name']] = key['field_value']
                # print(key)

            item.update(ticket_detail_format)
            export_list.append(item)

        # print(export_list)
        # 创建一个新的工作簿
        workbook = Workbook()
        # 获取工作表
        sheet = workbook.active
        # 写入表头
        header = list(export_list[0].keys())
        header[header.index('创建人')] = '提单人'
        header[header.index('creator')] = '创建人'
        header[header.index('gmt_created')] = '创建时间'
        header[header.index('card_bank')] = '结算银行'
        header[header.index('card_number')] = '结算卡号'
        header[header.index('account_balance_amount_before')] = '结算前余额'
        header[header.index('account_balance_amount_after')] = '结算后余额'
        header[header.index('total')] = '结算金额'
        header[header.index('capital_flow_type')] = '流水类型'
        header[header.index('ticket_record_id')] = '工单号'
        sheet.append(header)

        display_text = "点击查看附件"
        for index, item in enumerate(export_list):
            row = list(item.values())
            attachment_column = get_column_letter(len(row))
            attachment_link = row[-1]

            row = row[:-1]
            sheet.append(row)

            if attachment_link != "[]":
                add_hyperlink(sheet, sheet[f'{attachment_column}{index + 2}'], attachment_link, display_text)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=data.xlsx'
        workbook.save(response)

        # print(response)

        return response

