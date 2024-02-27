import json

import jwt
from django.contrib.auth.hashers import make_password
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect
from django.utils.decorators import method_decorator
from django.views import View
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from apps.account.models import LoonDept, LoonUser, LoonUserDept
from service.account.account_base_service import account_base_service_ins
from service.format_response import api_response
from apps.loon_base_view import LoonBaseView
from schema import Schema, Regex, And, Or, Use, Optional

from service.permission.manage_permission import manage_permission_check


@method_decorator(login_required, name='dispatch')
class LoonUserView(LoonBaseView):
    post_schema = Schema({
        'username': And(str, lambda n: n != '', error='username is needed'),
        'alias': And(str, lambda n: n != '', error='alias is needed'),
        'email': And(str, lambda n: n != '', error='email is needed'),
        Optional('password'): str,
        'phone': str,
        'dept_ids': str,
        'type_id': int,
        'is_active': Use(bool),

    })

    @manage_permission_check('workflow_admin')
    def get(self, request, *args, **kwargs):
        """
        获取用户列表
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        request_data = request.GET
        search_value = request_data.get('search_value', '')
        per_page = int(request_data.get('per_page', 10))
        page = int(request_data.get('page', 1))

        flag, result = account_base_service_ins.get_user_list(search_value, page, per_page)
        if flag is not False:
            data = dict(value=result.get('user_result_object_format_list'),
                        per_page=result.get('paginator_info').get('per_page'),
                        page=result.get('paginator_info').get('page'),
                        total=result.get('paginator_info').get('total'))
            code, msg,  = 0, ''
        else:
            code, data, msg = -1, '', result
        return api_response(code, msg, data)

    @manage_permission_check('admin')
    def post(self, request, *args, **kwargs):
        """
        add user
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        json_str = request.body.decode('utf-8')
        if not json_str:
            return api_response(-1, 'post参数为空', {})
        request_data_dict = json.loads(json_str)
        username = request_data_dict.get('username')
        alias = request_data_dict.get('alias')
        email = request_data_dict.get('email')
        password = request_data_dict.get('password')
        phone = request_data_dict.get('phone')
        dept_ids = request_data_dict.get('dept_ids')
        is_active = request_data_dict.get('is_active')
        type_id = request_data_dict.get('type_id')
        creator = request.user.username
        flag, result = account_base_service_ins.add_user(username, alias, email, phone, dept_ids, is_active, type_id, creator, password)
        if flag is False:
            code, msg, data = -1, result, {}
        else:
            code, msg, data = 0, '', result
        return api_response(code, msg, data)


@method_decorator(login_required, name='dispatch')
class LoonUserDetailView(LoonBaseView):
    patch_schema = Schema({
        'username': And(str, lambda n: n != ''),
        # 'alias': And(str, lambda n: n != ''),
        # 'email': And(str, lambda n: n != ''),
        Optional('password'): str,
        # 'phone': str,
        'dept_ids': str,
        'is_active': Use(bool),
        'type_id': int
    })

    @manage_permission_check('admin')
    def patch(self, request, *args, **kwargs):
        """
        edit user
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        json_str = request.body.decode('utf-8')
        user_id = kwargs.get('user_id')
        request_data_dict = json.loads(json_str)
        username = request_data_dict.get('username')
        alias = request_data_dict.get('alias')
        email = request_data_dict.get('email')
        phone = request_data_dict.get('phone')
        dept_ids = request_data_dict.get('dept_ids')
        type_id = request_data_dict.get('type_id')

        is_active = request_data_dict.get('is_active')
        flag, result = account_base_service_ins.edit_user(user_id, username, alias, email, phone, dept_ids, is_active,
                                                          type_id)
        if flag is not False:
            code, msg, data = 0, '', {}
        else:
            code, msg, data = -1, result, {}
        return api_response(code, msg, data)

    @manage_permission_check('admin')
    def delete(self, request, *args, **kwargs):
        """
        delete user record
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        user_id = kwargs.get('user_id')
        flag, result = account_base_service_ins.delete_user(user_id)
        if flag:
            code, msg, data = 0, '', {}
            return api_response(code, msg, data)
        code, msg, data = -1, result, {}
        return api_response(code, msg, data)


@method_decorator(login_required, name='dispatch')
class LoonRoleView(LoonBaseView):
    post_schema = Schema({
        'name': And(str, lambda n: n != ''),
        Optional('description'): str,
        Optional('label'): str,
    })

    @manage_permission_check('admin')
    def get(self, request, *args, **kwargs):
        """
        用户角色列表
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        request_data = request.GET
        search_value = request_data.get('search_value', '')
        per_page = int(request_data.get('per_page', 10))
        page = int(request_data.get('page', 1))
        flag, result = account_base_service_ins.get_role_list(search_value, page, per_page)
        if flag is not False:
            data = dict(value=result.get('role_result_object_format_list'),
                        per_page=result.get('paginator_info').get('per_page'),
                        page=result.get('paginator_info').get('page'),
                        total=result.get('paginator_info').get('total'))
            code, msg, = 0, ''
        else:
            code, data = -1, ''
        return api_response(code, msg, data)

    @manage_permission_check('admin')
    def post(self, request, *args, **kwargs):
        """
        add role
        新增角色
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        json_str = request.body.decode('utf-8')
        request_data_dict = json.loads(json_str)
        name = request_data_dict.get('name')
        description = request_data_dict.get('description', '')
        label = request_data_dict.get('label', '')
        creator = request.user.username

        flag, result = account_base_service_ins.add_role(name=name, description=description, label=label,
                                                         creator=creator)
        if flag is False:
            return api_response(-1, result, {})
        return api_response(0, result, {})


@method_decorator(login_required, name='dispatch')
class LoonRoleDetailView(LoonBaseView):
    patch_schema = Schema({
        'name': And(str, lambda n: n != '', error='name is need'),
        Optional('description'): str,
        Optional('label'): str,
    })

    @manage_permission_check('admin')
    def patch(self, request, *args, **kwargs):
        """
        update role
        更新角色信息
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        role_id = kwargs.get('role_id')
        json_str = request.body.decode('utf-8')
        request_data_dict = json.loads(json_str)
        name = request_data_dict.get('name')
        description = request_data_dict.get('description')
        label = request_data_dict.get('label')
        flag, result = account_base_service_ins.update_role(role_id, name, description, label)
        if flag is False:
            return api_response(-1, result, {})
        return api_response(0, '', {})

    @manage_permission_check('admin')
    def delete(self, request, *args, **kwargs):
        """
        delete role
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        role_id = kwargs.get('role_id')
        flag, result = account_base_service_ins.delete_role(role_id)
        if flag is False:
            return api_response(-1, result, {})
        return api_response(0, '', {})


@method_decorator(login_required, name='dispatch')
class LoonDeptView(LoonBaseView):
    post_schema = Schema({
        'name': And(str, lambda n: n != ''),
        Optional('parent_dept_id'): int,
        Optional('leader'): str,
        Optional('approver'): str,
        Optional('label'): str,
    })

    @manage_permission_check('admin')
    def get(self, request, *args, **kwargs):
        """
        部门列表
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        request_data = request.GET
        search_value = request_data.get('search_value', '')
        per_page = int(request_data.get('per_page', 10))
        page = int(request_data.get('page', 1))
        flag, result = account_base_service_ins.get_dept_list(search_value, page, per_page)
        if flag is not False:
            paginator_info = result.get('paginator_info')
            data = dict(value=result.get('dept_result_object_format_list'), per_page=paginator_info.get('per_page'),
                        page=paginator_info.get('page'), total=paginator_info.get('total'))
            code, msg, = 0, ''
        else:
            code, data = -1, ''
        return api_response(code, msg, data)

    @manage_permission_check('admin')
    def post(self, request, *args, **kwargs):
        """
        新增部门
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        json_str = request.body.decode('utf-8')
        request_data_dict = json.loads(json_str)
        name = request_data_dict.get('name')
        parent_dept_id = request_data_dict.get('parent_dept_id')
        leader = request_data_dict.get('leader')
        approver = request_data_dict.get('approver')

        label = request_data_dict.get('label')
        creator = request.user.username
        flag, result = account_base_service_ins.add_dept(name, parent_dept_id, leader, approver, label, creator)
        if flag is False:
            return api_response(-1, result, {})
        return api_response(0, result, {})


@method_decorator(login_required, name='dispatch')
class LoonDeptImportTemplateView(LoonBaseView):
    @manage_permission_check('admin')
    def get(self, request, *args, **kwargs):
        workbook = Workbook()
        work_sheet = workbook.active

        # 创建表头
        # field_items = serializer_class().get_fields().items()
        field_items = LoonDept._meta.fields
        field_names = [field.verbose_name for field in field_items if field.name !="id" and field.name !="creator" and field.name !="gmt_created" and field.name !="gmt_modified" and field.name !="is_deleted"]
        field_names[field_names.index('上级部门id')] = '上级部门'

        for index, field in enumerate(field_names, start=1):
            work_sheet.cell(row=1, column=index, value=field)
        # print(field_names)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=data.xlsx'
        workbook.save(response)

        return response

@method_decorator(login_required, name='dispatch')
class LoonDeptImportDataView(LoonBaseView):
    @manage_permission_check('admin')
    @transaction.atomic
    def post(self, request, *args, **kwargs):
        data = request.FILES['file']
        workbook = load_workbook(data)
        work_sheet = workbook.active

        field_items = LoonDept._meta.fields
        field_verbose_name = [field.verbose_name for field in field_items if
                       field.name != "id" and field.name != "creator" and field.name != "gmt_created" and field.name != "gmt_modified" and field.name != "is_deleted"]
        field_name = [field.name for field in field_items if
                              field.name != "id" and field.name != "creator" and field.name != "gmt_created" and field.name != "gmt_modified" and field.name != "is_deleted"]

        field_verbose_name[field_verbose_name.index('上级部门id')] = '上级部门'

        for index, field in enumerate(field_verbose_name):
            if work_sheet[1][index].value != field:
                code, msg, data = -1, '格式错误', {}
                return api_response(code, msg, data)

        data = []
        for row in range(2, work_sheet.max_row + 1):
            instance_item = {}
            for column, field in enumerate(field_name):
                if field == "parent_dept_id":
                    if work_sheet[row][column].value is not None and work_sheet[row][column].value != '':
                        try:
                            instance_item[field] = LoonDept.objects.filter(name=work_sheet[row][column].value).first().id
                        except Exception as e:
                            return api_response(-1, '上级部门填写有误！', {})
                    else:
                        instance_item[field] = 0
                else:
                    if work_sheet[row][column].value is not None:
                        instance_item[field] = work_sheet[row][column].value
            else:
                data.append(instance_item)

        # print(data)

        dept_names = {item['name'] for item in data}
        dept_set = LoonDept.objects.filter(name__in=dept_names)

        create_dept_set = []
        update_dept_set = []
        for item in data:
            for dept in dept_set:
                if dept.name == item['name']:
                    update_dept_set.append(dept)
                    for key, value in item.items():
                        setattr(dept, key, value)
                    break
            else:
                create_dept_set.append(LoonDept(**item))
        else:
            try:
                LoonDept.objects.bulk_create(create_dept_set)
                LoonDept.objects.bulk_update(update_dept_set,
                                                  field_name)
            except Exception as e:
                return api_response(-1, e.__str__(), {})

        return api_response(0, '', {})


@method_decorator(login_required, name='dispatch')
class LoonDeptExportView(LoonBaseView):
    @manage_permission_check('admin')
    def get(self, request, *args, **kwargs):
        request.GET = request.GET.copy()
        request.GET['per_page']= 99
        # request.GET['page']= 1
        username = request.META.get('HTTP_USERNAME')
        result = LoonDeptView.get(self, request, *args, **kwargs)

        # 将字节数据转换为字符串
        data_str = result.content.decode('utf-8')
        # 解析JSON字符串
        json_data = json.loads(data_str)

        export_list = []

        for item in json_data['data']['value']:
            export_dict = dict()
            export_dict['名称'] = item['name']

            if item['parent_dept_id'] != 0:
                export_dict['上级部门'] = LoonDept.objects.filter(id=item['parent_dept_id']).first().name
            else:
                export_dict['上级部门'] = ''

            export_dict['部门leader'] = item['leader']
            export_dict['审批人'] = item['approver']
            export_dict['标签'] = item['label']
            export_list.append(export_dict)

            # print(item)
            # 创建一个新的工作簿
        workbook = Workbook()
            # 获取工作表
        sheet = workbook.active
        header = ['名称','上级部门','部门leader','审批人','标签']
        sheet.append(header)

        for item in export_list:
            row = list(item.values())
            sheet.append(row)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=data.xlsx'
        workbook.save(response)

        return response



@method_decorator(login_required, name='dispatch')
class LoonDeptDetailView(LoonBaseView):
    patch_schema = Schema({
        'name': And(str, lambda n: n != '', error='name is need'),
        Optional('parent_dept_id'): int,
        Optional('leader'): str,
        Optional('approver'): str,
        Optional('label'): str,
    })

    @manage_permission_check('admin')
    def delete(self, request, *args, **kwargs):
        """
        delete dept
        删除部门
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        operator = request.user.username
        dept_id = kwargs.get('dept_id')
        flag, result = account_base_service_ins.delete_dept(dept_id)
        if flag is False:
            return api_response(-1, result, {})
        return api_response(0, '', {})

    @manage_permission_check('admin')
    def patch(self, request, *args, **kwargs):
        """
        更新部门
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        dept_id = kwargs.get('dept_id')
        json_str = request.body.decode('utf-8')
        request_data_dict = json.loads(json_str)
        name = request_data_dict.get('name')
        parent_dept_id = request_data_dict.get('parent_dept_id')
        leader = request_data_dict.get('leader')
        approver = request_data_dict.get('approver')
        label = request_data_dict.get('label')

        flag, result = account_base_service_ins.update_dept(dept_id,name, parent_dept_id, leader,
                                                            approver, label)
        if flag is False:
            return api_response(-1, result, {})
        return api_response(0, '', {})


class LoonSimpleDeptView(LoonBaseView):
    def get(self, request, *args, **kwargs):
        """
        部门列表，简单信息
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        request_data = request.GET
        search_value = request_data.get('search_value', '')
        per_page = int(request_data.get('per_page', 10))
        page = int(request_data.get('page', 1))
        flag, result = account_base_service_ins.get_dept_list(search_value, page, per_page, simple=True)
        if flag is not False:
            paginator_info = result.get('paginator_info')
            data = dict(value=result.get('dept_result_object_format_list'), per_page=paginator_info.get('per_page'),
                        page=paginator_info.get('page'), total=paginator_info.get('total'))
            code, msg, = 0, ''
        else:
            code, data = -1, ''
        return api_response(code, msg, data)


@method_decorator(login_required, name='dispatch')
class LoonAppTokenView(LoonBaseView):
    post_schema = Schema({
        'app_name': And(str, lambda n: n != '', error='app_name is needed'),
        Optional('ticket_sn_prefix'): str,
        Optional('workflow_ids'): str,
    })

    @manage_permission_check('admin')
    def get(self, request, *args, **kwargs):
        """
        call api permission
        调用权限列表
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        request_data = request.GET
        search_value = request_data.get('search_value', '')
        per_page = int(request_data.get('per_page', 10))
        page = int(request_data.get('page', 1))
        flag, result = account_base_service_ins.get_token_list(search_value, page, per_page)
        if flag is not False:
            paginator_info = result.get('paginator_info')
            data = dict(value=result.get('token_result_object_format_list'), per_page=paginator_info.get('per_page'),
                        page=paginator_info.get('page'), total=paginator_info.get('total'))
            code, msg, = 0, ''
        else:
            code, data, msg = -1, '', result
        return api_response(code, msg, data)

    @manage_permission_check('admin')
    def post(self, request, *args, **kwargs):
        """
        add call api permission
        新增调用权限记录
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        json_str = request.body.decode('utf-8')
        request_data_dict = json.loads(json_str)
        app_name = request_data_dict.get('app_name', '')
        ticket_sn_prefix = request_data_dict.get('ticket_sn_prefix', '')
        workflow_ids = request_data_dict.get('workflow_ids', '')
        username = request.user.username
        flag, result = account_base_service_ins.add_token_record(app_name, ticket_sn_prefix, workflow_ids, username)

        if flag is False:
            code, data = -1, {}
        else:
            code, data = 0, {'id': result.get('app_token_id')}

        return api_response(code, result, data)


@method_decorator(login_required, name='dispatch')
class LoonSimpleAppTokenView(LoonBaseView):
    @manage_permission_check('workflow_admin')
    def get(self, request, *args, **kwargs):
        """
        call api permission
        调用权限列表（返回简单数据）
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        request_data = request.GET
        search_value = request_data.get('search_value', '')
        per_page = int(request_data.get('per_page', 10))
        page = int(request_data.get('page', 1))
        flag, result = account_base_service_ins.get_token_list(search_value, page, per_page, simple=True)
        if flag is not False:
            paginator_info = result.get('paginator_info')
            data = dict(value=result.get('token_result_object_format_list'), per_page=paginator_info.get('per_page'),
                        page=paginator_info.get('page'), total=paginator_info.get('total'))
            code, msg, = 0, ''
        else:
            code, data, msg = -1, '', result
        return api_response(code, msg, data)



@method_decorator(login_required, name='dispatch')
class LoonAppTokenDetailView(LoonBaseView):
    patch_schema = Schema({
        Optional('ticket_sn_prefix'): str,
        Optional('workflow_ids'): str,
    })

    @manage_permission_check('admin')
    def patch(self, request, *args, **kwargs):
        """
        编辑token
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        app_token_id = kwargs.get('app_token_id')
        json_str = request.body.decode('utf-8')
        request_data_dict = json.loads(json_str)
        ticket_sn_prefix = request_data_dict.get('ticket_sn_prefix', '')
        workflow_ids = request_data_dict.get('workflow_ids', '')
        flag, msg = account_base_service_ins.update_token_record(app_token_id, ticket_sn_prefix, workflow_ids)
        if flag is False:
            code, data = -1, {}
        else:
            code, data = 0, {}

        return api_response(code, msg, data)

    @manage_permission_check('admin')
    def delete(self, request, *args, **kwargs):
        """
        删除记录
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        app_token_id = kwargs.get('app_token_id')
        flag, msg = account_base_service_ins.del_token_record(app_token_id)
        if flag is False:
            code, data = -1, {}
        else:
            code, data = 0, {}
        return api_response(code, msg, data)


class LoonLoginView(LoonBaseView):
    def post(self, request, *args, **kwargs):
        """
        登录验证
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        json_str = request.body.decode('utf-8')
        if not json_str:
            return api_response(-1, 'patch参数为空', {})
        request_data_dict = json.loads(json_str)
        username = request_data_dict.get('username', '')
        password = request_data_dict.get('password', '')
        try:
            user = authenticate(username=username, password=password)
        except Exception as e:
            return api_response(-1, e.__str__(), {})

        if user is not None:
            login(request, user)
            return api_response(0, '', {})
        else:
            return api_response(-1, 'username or password is invalid', {})


class LoonLogoutView(LoonBaseView):
    def get(self, request, *args, **kwargs):
        """
        注销
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        logout(request)
        return redirect('/manage')


class LoonJwtLoginView(LoonBaseView):
    def post(self, request, *args, **kwargs):
        json_str = request.body.decode('utf-8')
        if not json_str:
            return api_response(-1, 'invalid args', {})
        request_data_dict = json.loads(json_str)
        username = request_data_dict.get('username', '')
        password = request_data_dict.get('password', '')
        try:
            user = authenticate(username=username, password=password)
        except Exception as e:
            return api_response(-1, e.__str__(), {})
        if user is not None:
            # todo: get jwt
            flag, jwt_info = account_base_service_ins.get_user_jwt(username)
            if flag is False:
                return api_response(-1, '', {})
            else:
                login(request, user)
                return api_response(0, '', {'jwt': jwt_info})
        else:
            return api_response(-1, 'username or password is invalid', {})


@method_decorator(login_required, name='dispatch')
class LoonUserRoleView(LoonBaseView):
    @manage_permission_check('admin')
    def get(self, request, *args, **kwargs):
        """
        用户角色信息
        """
        user_id = kwargs.get('user_id', 0)
        search_value = request.GET.get('search_value', '')
        flag, result = account_base_service_ins.get_user_role_info_by_user_id(user_id, search_value)
        if flag is not False:
            data = dict(value=result.get('role_result_format_list'), per_page=result.get('paginator_info').get('per_page'),
                        page=result.get('paginator_info').get('page'), total=result.get('paginator_info').get('total'))
            code, msg, = 0, ''
        else:
            code, data = -1, ''
        return api_response(code, msg, data)


@method_decorator(login_required, name='dispatch')
class LoonRoleUserView(LoonBaseView):
    post_schema = Schema({
        'user_id': And(int, error='user_id is needed and should be int')
    })

    @manage_permission_check('admin')
    def get(self, request, *args, **kwargs):
        """
        角色的用户信息
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        role_id = kwargs.get('role_id', 0)
        request_data = request.GET
        page = int(request_data.get('page', 1))
        per_page = int(request_data.get('per_page', 10))
        search_value = request.GET.get('search_value', '')
        flag, result = account_base_service_ins.get_role_user_info_by_role_id(role_id, search_value, page, per_page)

        if flag is not False:
            data = dict(value=result.get('user_result_format_list'), per_page=result.get('paginator_info').get('per_page'),
                        page=result.get('paginator_info').get('page'), total=result.get('paginator_info').get('total'))
            code, msg, = 0, ''
        else:
            code, data = -1, ''
        return api_response(code, msg, data)

    @manage_permission_check('admin')
    def post(self, request, *args, **kwargs):
        """
        add role's user
        新增角色用户
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        role_id = kwargs.get('role_id', 0)
        creator = request.user.username
        json_str = request.body.decode('utf-8')
        request_data_dict = json.loads(json_str)
        user_id = request_data_dict.get('user_id', 0)

        flag, result = account_base_service_ins.add_role_user(role_id, user_id, creator)
        if flag is False:
            return api_response(-1, result, {})
        return api_response(0, '', {})


@method_decorator(login_required, name='dispatch')
class LoonRoleUserDetailView(LoonBaseView):
    @manage_permission_check('admin')
    def delete(self, request, *args, **kwargs):
        """
         delete role's user
         删除角色用户
         :param request:
         :param args:
         :param kwargs:
         :return:
         """
        user_id = kwargs.get('user_id', 0)
        flag, result = account_base_service_ins.delete_role_user(user_id)

        if flag is False:
            return api_response(-1, result, {})
        return api_response(0, '', {})


class LoonUserResetPasswordView(LoonBaseView):

    @manage_permission_check('admin')
    def post(self, request, *args, **kwargs):
        """
        重置密码
        :param requesdt:
        :param args:
        :param kwargs:
        :return:
        """
        user_id = kwargs.get('user_id')
        flag, result = account_base_service_ins.reset_password(user_id=user_id)
        if flag is False:
            return api_response(-1, result, {})
        return api_response(0, result, {})


class LoonUserChangePasswordView(LoonBaseView):
    def post(self, request, *args, **kwargs):
        """
        修改密码
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        username = request.user.username

        json_str = request.body.decode('utf-8')
        request_data_dict = json.loads(json_str)
        new_password = request_data_dict.get('new_password', '')
        source_password = request_data_dict.get('source_password', '')
        new_password_again = request_data_dict.get('new_password_again', '')

        if new_password != new_password_again:
            return api_response(-1, '两次密码不一致，请重新输入', {})
        flag, result = account_base_service_ins.change_password(username, source_password, new_password)
        if flag is False:
            return api_response(-1, result, {})
        return api_response(0, result, {})

class LoonUserImportTemplateView(LoonBaseView):
    @manage_permission_check('admin')
    def get(self, request, *args, **kwargs):
        workbook = Workbook()
        work_sheet = workbook.active

        field_names = ['用户名(必填)', '部门(必填,以\",\"分割)', '状态(在职/离职)(必填)', '用户类型(普通用户/工作流管理员/财务管理员/超级管理员)(必填)']

        for index, field in enumerate(field_names, start=1):
            work_sheet.cell(row=1, column=index, value=field)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=data.xlsx'
        workbook.save(response)

        return response


class LoonUserImportDataTemplateView(LoonBaseView):
    @manage_permission_check('admin')
    @transaction.atomic
    def post(self, request, *args, **kwargs):
        data = request.FILES['file']
        workbook = load_workbook(data)
        work_sheet = workbook.active

        field_names = ['用户名(必填)', '部门(必填,以\",\"分割)', '状态(在职/离职)(必填)', '用户类型(普通用户/工作流管理员/财务管理员/超级管理员)(必填)']

        for index, field in enumerate(field_names):
            if work_sheet[1][index].value != field:
                code, msg, data = -1, '格式错误', {}
                return api_response(code, msg, data)

        data = []
        for row in range(2, work_sheet.max_row + 1):
            instance_item = {}
            for column, field in enumerate(field_names):
                if work_sheet[row][column].value is None or work_sheet[row][column].value == '':
                    return api_response(-1, '缺少必填项', {})
                else:
                    if field == "用户名(必填)":
                        instance_item["username"] = work_sheet[row][column].value
                    elif field == "部门(必填,以\",\"分割)":
                        try:
                            dept_ids = []
                            depts = work_sheet[row][column].value.split(',')
                            for dept in depts:
                                dept_ids.append(LoonDept.objects.filter(name=dept).first().id)
                            instance_item["dept_ids"] = dept_ids
                        except Exception as e:
                            return api_response(-1, '部门填写有误！', {})
                    elif field == "状态(在职/离职)(必填)":
                        if work_sheet[row][column].value not in ['在职', '离职']:
                            return api_response(-1, '状态填写有误！', {})
                        else:
                            if work_sheet[row][column].value == '在职':
                                instance_item['is_active'] = 1
                            else:
                                instance_item['is_active'] = 0
                    elif field == "用户类型(普通用户/工作流管理员/财务管理员/超级管理员)(必填)":
                        if work_sheet[row][column].value not in ['普通用户', '工作流管理员', '财务管理员', '超级管理员']:
                            return api_response(-1, '用户类型填写有误！', {})
                        else:
                            if work_sheet[row][column].value == '普通用户':
                                instance_item['type_id'] = 0
                            elif work_sheet[row][column].value == '工作流管理员':
                                instance_item['type_id'] = 1
                            elif work_sheet[row][column].value == '超级管理员':
                                instance_item['type_id'] = 2
                            else:
                                instance_item['type_id'] = 3
            else:
                data.append(instance_item)

        # print(data)

        user_names = {item['username'] for item in data}
        user_set = LoonUser.objects.filter(username__in=user_names)


        update_user_set = []
        for item in data:
            for user in user_set:
                if user.username == item['username']:
                    update_user_set.append(user)
                    for key, value in item.items():
                        if key == "dept_ids":
                            user_dept_ids_list_raw = [ user.dept_id for user in LoonUserDept.objects.filter(user_id=user.id, is_deleted=0).all() ]
                            add_ids = list(set(value) - set(user_dept_ids_list_raw))
                            delete_ids = list(set(user_dept_ids_list_raw) - set(value))

                            queryset_add_list = []
                            for dept_id in add_ids:
                                queryset_add_list.append(LoonUserDept(user_id=user.id, dept_id=dept_id))
                            LoonUserDept.objects.bulk_create(queryset_add_list)

                            for dept_id in delete_ids:
                                LoonUserDept.objects.filter(user_id=user.id, dept_id=dept_id).update(is_deleted=1)
                            continue
                        setattr(user, key, value)
                    break
            else:
                password_str = make_password("123456", None, 'pbkdf2_sha256')
                user_obj = LoonUser(username=item['username'],
                                    is_active=item['is_active'], type_id=item['type_id'],
                                    creator="admin", password=password_str)
                user_obj.save()

                dept_ids = ','.join(str(id) for id in item.pop('dept_ids', None))
                queryset_list = []
                for dept_id in dept_ids.split(','):
                    queryset_list.append(LoonUserDept(user_id=user_obj.id, dept_id=dept_id))
                LoonUserDept.objects.bulk_create(queryset_list)
        else:
            try:
                LoonUser.objects.bulk_update(update_user_set,
                                             ['is_active','type_id'])
            except Exception as e:
                return api_response(-1, e.__str__(), {})

        return api_response(0, '', {})


class LoonUserExportView(LoonBaseView):
    @manage_permission_check('admin')
    def get(self, request, *args, **kwargs):
        request.GET = request.GET.copy()
        request.GET['per_page'] = 999
        # request.GET['page']= 1
        username = request.META.get('HTTP_USERNAME')
        result = LoonUserView.get(self, request, *args, **kwargs)

        # 将字节数据转换为字符串
        data_str = result.content.decode('utf-8')
        # 解析JSON字符串
        json_data = json.loads(data_str)

        export_list = []

        type_dict = {'0':'普通用户', '1':'工作流管理员', '2':'超级管理员', '3':'财务管理员'}

        for item in json_data['data']['value']:
            export_dict = dict()
            export_dict['用户名(必填)'] = item['username']
            export_dict['部门(必填,以","分割)'] = ','.join(str(dept) for dept in [ dept['name'] for dept in item['user_dept_info_list'] ])
            export_dict['状态(在职/离职)(必填)'] = '在职' if item['is_active'] else '离职'
            export_dict['用户类型(普通用户/工作流管理员/财务管理员/超级管理员)(必填)'] = type_dict.get(str(item['type_id']))

            export_list.append(export_dict)

        workbook = Workbook()
        sheet = workbook.active
        header = ['用户名(必填)', '部门(必填,以","分割)', '状态(在职/离职)(必填)', '用户类型(普通用户/工作流管理员/财务管理员/超级管理员)(必填)']
        sheet.append(header)

        for item in export_list:
            row = list(item.values())
            sheet.append(row)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment;filename=data.xlsx'
        workbook.save(response)

        return response



class LoonSimpleUserView(LoonBaseView):

    def get(self, request, *args, **kwargs):
        """
        获取用户简要信息列表
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        request_data = request.GET
        search_value = request_data.get('search_value', '')
        per_page = int(request_data.get('per_page', 10))
        page = int(request_data.get('page', 1))
        flag, result = account_base_service_ins.get_user_list(search_value, page, per_page, simple=True)
        if flag is not False:
            data = dict(value=result.get('user_result_object_format_list'),
                        per_page=result.get('paginator_info').get('per_page'),
                        page=result.get('paginator_info').get('page'),
                        total=result.get('paginator_info').get('total'))
            code, msg, = 0, ''
        else:
            code, data, msg = -1, '', result
        return api_response(code, msg, data)
